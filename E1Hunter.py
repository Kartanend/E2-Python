# Autor: Diego Moreno Villarreal
from pyhunter import PyHunter
from openpyxl import Workbook
import getpass


# Cantidad de resultados esperados en la busqueda
# El límite MENSUAL de Hunter es 50, cuidado!
def busqueda(organizacion):
    resultado = hunter.domain_search(company=organizacion, limit=1,
                                     emails_type='personal')
    return resultado


def guardar_informacion(datosEncontrados, organizacion):
    libro = Workbook()
    hoja = libro.create_sheet(organizacion)
    libro.save("Hunter" + organizacion + ".xlsx")
    hoja.cell(1, 1, "Correo")
    hoja.cell(1, 2, "Nombre")
    hoja.cell(1, 3, "Apellido")
    fila = 2
    for usuario in datosEncontrados["emails"]:
        hoja.cell(fila, 1, usuario["value"])
        hoja.cell(fila, 2, usuario["first_name"])
        hoja.cell(fila, 3, usuario["last_name"])
        fila += 1
    libro.save("Hunter" + organizacion + ".xlsx")
    print("Los datos se han guardado satisfactoriamente!")


print("Script para buscar información")
apikey = getpass.getpass("Ingresa tu API key: ")
hunter = PyHunter(apikey)
orga = input("Dominio a investigar: ")
datosEncontrados = busqueda(orga)
if datosEncontrados is None:
    exit()
else:
    print(datosEncontrados)
    print(type(datosEncontrados))
    guardar_informacion(datosEncontrados, orga)
